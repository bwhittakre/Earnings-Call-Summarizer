#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel export helper for the Structured Narrative outputs.

Why this exists
---------------
CSV cannot carry cell number-formats or column widths, so when the output
sheets are opened in Excel the date/time columns (e.g. ``earnings_datetime``
= "2015-01-29 16:01:00") are auto-detected as dates and rendered as
``########`` whenever the default column is too narrow.

``write_excel`` produces a real ``.xlsx`` where every date / datetime column
carries an explicit number format (``yyyy-mm-dd`` or ``yyyy-mm-dd hh:mm``) and
every column is widened to fit its contents, so nothing ever collapses to
``########``. The header row is frozen for easy scanning.

Used by single_company_extractor.py and narrative_zscore.py. Run directly to
(re)generate workbooks from every parquet under output/{TICKER}/parquet/ and
output/cross_company/parquet/:

    python "Structured Narrative/excel_export.py"
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from output_paths import (
    company_artifact,
    cross_company_artifact,
    cross_company_layer,
    list_company_tickers,
    resolve_read_parquet_or_csv,
    ROOT,
)
from dimension_order import insert_dimension_group_header_rows, sort_panel_by_dimension  # noqa: E402

DATE_FMT = "yyyy-mm-dd"
DATETIME_FMT = "yyyy-mm-dd hh:mm"
PERCENT_FMT = "0.00%"
ZSCORE_FMT = "0.000"
MAX_WIDTH = 42
TEXT_MAX_WIDTH = 80
TEXT_COLUMN_HINTS = ("rationale", "excerpt", "claim", "desc", "label", "text")

NARRATIVE_ZSCORED_DROP_COLS = frozenset({
    "earnings_datetime",
    "model_date",
    "measure_desc",
    "actual_effectivedate",
    "alpha_spec_0_60_complete",
    "alpha_spec_60_90_complete",
    "alpha_spec_0_90_complete",
    "window_overlaps_next_earnings",
    "src_consensus_view",
    "src_actual_view",
    "src_return_view",
    "return_model",
})

PROPORTION_SOURCE_COLS = frozenset({
    "earnings_surprise_pct",
    "fwd_estimate_revision_pct",
})

LAYER_COLUMN_LABELS: dict[str, str] = {
    "ticker": "Ticker",
    "estpermid": "Est Perm ID",
    "isin": "ISIN",
    "barra_id": "Barra ID",
    "fiscal_period": "Fiscal Period",
    "fiscal_quarter_end": "Fiscal Quarter End",
    "earnings_date": "Earnings Date",
    "next_earnings_date": "Next Earnings Date",
    "measure": "Measure Code",
    "measure_label": "Measure",
    "period_role": "Period Role",
    "target_pertype": "Target Per Type",
    "target_period": "Target Period",
    "target_period_end": "Target Period End",
    "actual_value": "Actual Value",
    "consensus_pre_mean": "Consensus Pre Mean",
    "consensus_pre_median": "Consensus Pre Median",
    "consensus_pre_high": "Consensus Pre High",
    "consensus_pre_low": "Consensus Pre Low",
    "consensus_pre_numests": "Consensus Pre # Ests",
    "consensus_pre_effectivedate": "Consensus Pre Effective Date",
    "consensus_post7_mean": "Consensus Post-7d Mean",
    "consensus_post7_numests": "Consensus Post-7d # Ests",
    "consensus_post7_effectivedate": "Consensus Post-7d Effective Date",
    "earnings_surprise": "Earnings Surprise",
    "earnings_surprise_pct": "Earnings Surprise %",
    "fwd_estimate_revision": "Fwd Estimate Revision",
    "fwd_estimate_revision_pct": "Fwd Estimate Revision %",
    "unittype": "Unit Type",
    "defscale": "Def Scale",
    "alpha_spec_0_60": "Alpha Spec 0-60",
    "alpha_spec_60_90": "Alpha Spec 60-90",
    "alpha_spec_0_90": "Alpha Spec 0-90",
    "earnings_surprise_pct_z": "Earnings Surprise Z",
    "earnings_surprise_pct_z_pit": "Earnings Surprise Z (PIT)",
    "fwd_estimate_revision_pct_z": "Fwd Estimate Revision Z",
    "fwd_estimate_revision_pct_z_pit": "Fwd Estimate Revision Z (PIT)",
    "alpha_spec_0_90_z": "Alpha Spec 0-90 Z",
    "dimension": "Dimension",
    "quant_z": "Quant Z",
    "quant_z_pit": "Quant Z (PIT)",
    "as_of_date": "As Of Date",
    "is_quant_comparable": "Quant Comparable",
    "llm_level": "LLM Level",
    "surprise_direction": "Surprise Direction",
    "surprise_magnitude": "Surprise Magnitude",
    "agrees_with_quant": "Agrees With Quant",
    "narrative_quant_gap": "Narrative Quant Gap",
    "rationale": "Rationale",
    "n_evidence": "# Evidence",
    "n_evidence_verified": "# Evidence Verified",
    "evidence_verified": "Evidence Verified",
    "excerpts": "Excerpts",
    "source": "Source",
    "period_end_date": "Period End Date",
    "period_end_calendar_quarter": "Period End Calendar Qtr",
    "dimension_group": "Dimension Group",
    "call_feature_available_date": "Call Feature Available Date",
    "t7_feature_available_date": "T+7 Feature Available Date",
    "feature_availability_date": "Feature Available Date",
    "investable_as_of_date": "Investable As-Of Date",
    "days_since_earnings": "Days Since Earnings",
    "feature_age_days": "Feature Age (Days)",
    "investable_ready": "Investable Ready",
    "quant_mapping": "Quant Mapping",
    "quant_family": "Quant Family",
    "quant_z_fullsample": "Quant Z (Full Sample)",
    "quant_guidance_revision_z_pit": "Guidance Revision Z (T+7 PIT)",
    "alpha_spec_0_90_complete": "Alpha Spec 0-90 Complete",
    "alpha_spec_asof_0_90": "Alpha Spec As-Of 0-90",
    "alpha_spec_asof_0_90_z": "Alpha Spec As-Of 0-90 Z",
    "alpha_spec_asof_0_90_complete": "Alpha Spec As-Of 0-90 Complete",
    "in_cross_section": "In Cross Section",
    "exclusion_reason": "Exclusion Reason",
    "level_rationale": "Level Rationale",
    "prior_period": "Prior Period",
    "change_direction": "Change Direction",
    "change_magnitude": "Change Magnitude",
    "score_delta": "Score Delta",
    "quant_z_delta": "Quant Z Delta",
    "delta_rationale": "Delta Rationale",
    "level_evidence_supported_pct": "Level Evidence %",
    "delta_evidence_supported_pct": "Delta Evidence %",
    "surprise_evidence_supported_pct": "Surprise Evidence %",
    "novelty_direction": "Novelty Direction",
    "narrative_novelty": "Narrative Novelty",
    "novelty_rationale": "Novelty Rationale",
    "novelty_evidence_supported_pct": "Novelty Evidence %",
    "has_level": "Has Level",
    "has_delta": "Has Delta",
    "has_surprise": "Has Surprise",
    "has_novelty": "Has Novelty",
    "level_quant_sign_match": "Level Quant Sign Match",
    "delta_quant_sign_match": "Delta Quant Sign Match",
    "level_diverges": "Level Diverges",
    "delta_diverges": "Delta Diverges",
    "surprise_diverges": "Surprise Diverges",
    "any_quant_divergence": "Any Quant Divergence",
    "is_divergence": "Is Divergence",
    "has_quant_z": "Has Quant Z",
    "abs_narrative_quant_gap": "Abs Narrative Quant Gap",
    "surprise_quant_interaction": "Surprise Quant Interaction",
    "llm_level_4q_mean": "LLM Level 4Q Mean",
    "change_magnitude_4q_mean": "Change Magnitude 4Q Mean",
    "signal_stack": "Signal Stack",
    "evidence_confidence": "Evidence Confidence",
}

PANEL_EXCEL_FRONT_COLS = (
    "ticker",
    "fiscal_period",
    "period_end_date",
    "period_end_calendar_quarter",
    "dimension",
    "dimension_group",
    "as_of_date",
    "earnings_date",
    "call_feature_available_date",
    "t7_feature_available_date",
    "feature_availability_date",
    "investable_as_of_date",
    "days_since_earnings",
    "feature_age_days",
    "investable_ready",
    "quant_mapping",
    "quant_family",
    "llm_level",
    "change_direction",
    "change_magnitude",
    "surprise_direction",
    "surprise_magnitude",
    "narrative_novelty",
    "novelty_direction",
    "quant_z_pit",
    "quant_z",
    "quant_guidance_revision_z_pit",
    "agrees_with_quant",
    "narrative_quant_gap",
    "evidence_confidence",
)

EVIDENCE_PCT_COLS = frozenset({
    "level_evidence_supported_pct",
    "delta_evidence_supported_pct",
    "surprise_evidence_supported_pct",
    "novelty_evidence_supported_pct",
    "evidence_confidence",
})


def _humanize_column(name: str) -> str:
    if name in LAYER_COLUMN_LABELS:
        return LAYER_COLUMN_LABELS[name]
    return " ".join(part.capitalize() for part in name.split("_"))


def _is_zscore_source_col(name: str) -> bool:
    if name in PROPORTION_SOURCE_COLS:
        return False
    if name in ("quant_z", "quant_z_pit", "alpha_spec_0_90_z"):
        return True
    return name.endswith("_z") or name.endswith("_z_pit")


def _drop_layers_columns(df: pd.DataFrame, sheet: str) -> pd.DataFrame:
    if sheet == "narrative_zscored":
        drop = [c for c in NARRATIVE_ZSCORED_DROP_COLS if c in df.columns]
        if drop:
            return df.drop(columns=drop)
    return df


def _rename_layers_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, str]]:
    rename = {c: _humanize_column(c) for c in df.columns}
    return df.rename(columns=rename), rename


def _prepare_layers_sheet(
    df: pd.DataFrame,
    sheet: str,
) -> tuple[pd.DataFrame, dict[str, str], set[str], set[str]]:
    """Drop, rename, and classify columns for narrative_layers Excel export."""
    out = _drop_layers_columns(df, sheet)
    out, rename_map = _rename_layers_columns(out)
    percent_cols: set[str] = set()
    zscore_cols: set[str] = set()
    for src, label in rename_map.items():
        if src in PROPORTION_SOURCE_COLS:
            percent_cols.add(label)
        elif _is_zscore_source_col(src):
            zscore_cols.add(label)
    out, kinds = _prepare_sheet(out)
    return out, kinds, percent_cols, zscore_cols


def _date_kind(series: pd.Series, name: str):
    """Return 'date', 'datetime', or None for a column."""
    name_l = name.lower()
    looks_datey = any(k in name_l for k in ("date", "_end"))
    if not (pd.api.types.is_datetime64_any_dtype(series) or looks_datey):
        return None
    s = pd.to_datetime(series, errors="coerce")
    if s.notna().sum() == 0:
        return None
    nonnull = s.dropna()
    all_midnight = bool(
        (nonnull.dt.hour == 0).all()
        and (nonnull.dt.minute == 0).all()
        and (nonnull.dt.second == 0).all()
    )
    return "date" if all_midnight else "datetime"


def _is_text_column(name: str, series: pd.Series) -> bool:
    name_l = name.lower()
    if any(h in name_l for h in TEXT_COLUMN_HINTS):
        return True
    if pd.api.types.is_string_dtype(series) or series.dtype == object:
        sample = series.dropna().head(20)
        if sample.empty:
            return False
        return sample.map(lambda v: isinstance(v, str) and len(str(v)) > 24).any()
    return False


def _col_width(
    series: pd.Series,
    header: str,
    kind: str | None,
    *,
    full_scan: bool = False,
) -> float:
    if kind == "date":
        content = 10
        cap = MAX_WIDTH
    elif kind == "datetime":
        content = 16
        cap = MAX_WIDTH
    else:
        sample = series.dropna().astype(str)
        if not full_scan and len(sample) > 500:
            sample = sample.sample(500, random_state=0)
        content = int(sample.map(len).max()) if not sample.empty else 0
        cap = TEXT_MAX_WIDTH if _is_text_column(header, series) else MAX_WIDTH
    return min(max(content, len(str(header))) + 2, cap)


def _prepare_sheet(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, str]]:
    out = df.copy()
    kinds: dict[str, str] = {}
    for col in out.columns:
        kind = _date_kind(out[col], col)
        if kind:
            out[col] = pd.to_datetime(out[col], errors="coerce")
            kinds[col] = kind
    return out, kinds


def _table_display_name(sheet_name: str) -> str:
    safe = re.sub(r"[^\w]", "_", sheet_name)
    if not safe or safe[0].isdigit():
        safe = f"tbl_{safe}"
    return f"tbl_{safe}"[:255]


def _apply_excel_table(ws, *, nrows: int, ncols: int, display_name: str) -> None:
    if nrows < 1 or ncols < 1:
        return
    ref = f"A1:{get_column_letter(ncols)}{nrows}"
    tab = Table(displayName=display_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def _apply_group_header_rows(
    ws,
    header_row_indices: list[int],
    ncols: int,
    *,
    label_col: int = 1,
) -> None:
    """Merge and style inserted dimension-group label rows (0-based data frame indices)."""
    if not header_row_indices or ncols < 1:
        return
    last_col = get_column_letter(ncols)
    fill = PatternFill(fill_type="solid", fgColor="EEF2FF")
    font = Font(bold=True, color="1E3A8A")
    for idx in header_row_indices:
        excel_row = idx + 2
        label = ws.cell(row=excel_row, column=label_col).value
        ws.merge_cells(f"A{excel_row}:{last_col}{excel_row}")
        cell = ws.cell(row=excel_row, column=1)
        cell.value = label
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")


def _format_worksheet(
    ws,
    out: pd.DataFrame,
    kinds: dict[str, str],
    *,
    as_table: bool = False,
    table_name: str | None = None,
    full_scan: bool = False,
    percent_cols: set[str] | None = None,
    zscore_cols: set[str] | None = None,
    group_header_rows: list[int] | None = None,
) -> None:
    ws.freeze_panes = "A2"
    nrows = len(out) + 1
    ncols = len(out.columns)
    percent_cols = percent_cols or set()
    zscore_cols = zscore_cols or set()

    for i, col in enumerate(out.columns, start=1):
        letter = get_column_letter(i)
        kind = kinds.get(col)
        if kind:
            fmt = DATE_FMT if kind == "date" else DATETIME_FMT
            for cell in ws[letter][1:]:
                cell.number_format = fmt
        elif col in percent_cols:
            for cell in ws[letter][1:]:
                cell.number_format = PERCENT_FMT
        elif col in zscore_cols:
            for cell in ws[letter][1:]:
                cell.number_format = ZSCORE_FMT
        ws.column_dimensions[letter].width = _col_width(
            out[col], col, kind, full_scan=full_scan
        )

    group_header_rows = group_header_rows or []
    label_col = 1
    if "Dimension" in out.columns:
        label_col = out.columns.get_loc("Dimension") + 1
    if group_header_rows:
        _apply_group_header_rows(
            ws, group_header_rows, ncols, label_col=label_col
        )

    if as_table and not group_header_rows:
        _apply_excel_table(
            ws,
            nrows=nrows,
            ncols=ncols,
            display_name=table_name or _table_display_name(ws.title),
        )
    elif nrows > 1 and ncols >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows}"


def _dimension_z_columns(df: pd.DataFrame) -> list[str]:
    return sorted(
        c for c in df.columns
        if c.startswith("dim_") and c.endswith("_z") and not c.endswith("_z_pit")
    )


def _dimension_scores_for_excel(df: pd.DataFrame, ticker: str) -> pd.DataFrame:
    """Melt wide dimension_scores into long rows for filterable Excel tables."""
    t = ticker.upper()
    z_cols = _dimension_z_columns(df)
    if not z_cols:
        out = df.copy()
        if "ticker" not in out.columns:
            out.insert(0, "ticker", t)
        return out

    id_cols = [
        c for c in (
            "fiscal_period",
            "earnings_date",
            "alpha_spec_0_90",
            "alpha_spec_0_90_z",
        )
        if c in df.columns
    ]
    melted = df[id_cols + z_cols].melt(
        id_vars=id_cols,
        value_vars=z_cols,
        var_name="_dim_col",
        value_name="quant_z",
    )
    melted["dimension"] = melted["_dim_col"].str.removeprefix("dim_").str.removesuffix("_z")
    pit_cols = {c: c.replace("_z", "_z_pit") for c in z_cols}
    pit_lookup = df[id_cols + list(pit_cols.values())].copy()
    pit_lookup = pit_lookup.rename(columns={v: k for k, v in pit_cols.items()})
    pit_long = pit_lookup.melt(
        id_vars=id_cols,
        value_vars=z_cols,
        var_name="_dim_col",
        value_name="quant_z_pit",
    )
    pit_long["dimension"] = pit_long["_dim_col"].str.removeprefix("dim_").str.removesuffix("_z")
    merged = melted.merge(
        pit_long[id_cols + ["dimension", "quant_z_pit"]],
        on=id_cols + ["dimension"],
        how="left",
    ).drop(columns=["_dim_col"])

    merged.insert(0, "ticker", t)
    sort_cols = [c for c in ("fiscal_period", "dimension") if c in merged.columns]
    if sort_cols:
        merged = merged.sort_values(sort_cols).reset_index(drop=True)

    front = ["ticker", "fiscal_period", "dimension"]
    rest = [c for c in merged.columns if c not in front]
    return merged[front + rest]


def _rename_dataframe_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, str]]:
    rename = {c: _humanize_column(c) for c in df.columns}
    return df.rename(columns=rename), rename


def _sort_cross_section(df: pd.DataFrame, dimension_order: str | None = None) -> pd.DataFrame:
    return sort_panel_by_dimension(df, dimension_order)


def prepare_panel_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Reorder panel columns: identifiers and scores first, rationales last."""
    rationale_cols = [c for c in df.columns if "rationale" in c]
    front = [c for c in PANEL_EXCEL_FRONT_COLS if c in df.columns]
    middle = [
        c for c in df.columns
        if c not in front and c not in rationale_cols
    ]
    tail = [c for c in rationale_cols if c in df.columns]
    return df[front + middle + tail]


def _classify_renamed_columns(rename_map: dict[str, str]) -> tuple[set[str], set[str]]:
    percent_cols: set[str] = set()
    zscore_cols: set[str] = set()
    for src, label in rename_map.items():
        if src in EVIDENCE_PCT_COLS:
            percent_cols.add(label)
        elif _is_zscore_source_col(src):
            zscore_cols.add(label)
    return percent_cols, zscore_cols


def _prepare_cross_section_sheet(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, str], set[str], set[str]]:
    out = df.copy()
    out, rename_map = _rename_dataframe_columns(out)
    percent_cols, zscore_cols = _classify_renamed_columns(rename_map)
    out, kinds = _prepare_sheet(out)
    return out, kinds, percent_cols, zscore_cols


def write_cross_section_panel_workbook(
    path: str | Path,
    panel_df: pd.DataFrame,
    spine_df: pd.DataFrame,
    *,
    dimension_order: str | None = None,
) -> str:
    """Write filterable cross-section workbook: Summary (spine) + Panel (full rows).

    Group header rows stay on the formatted Summary sheet only. The Panel sheet
    is observation-only (use dimension_group for classification).
    """
    path = str(path)
    panel_sorted = _sort_cross_section(prepare_panel_for_excel(panel_df), dimension_order)
    spine_sorted = _sort_cross_section(spine_df, dimension_order)
    # Machine-readable Panel: no synthetic Fundamentals / Narrative Context rows.
    panel_export, panel_header_rows = panel_sorted, []
    spine_export, spine_header_rows = insert_dimension_group_header_rows(spine_sorted)

    sheets = {
        "Summary": (spine_export, spine_header_rows),
        "Panel": (panel_export, panel_header_rows),
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, (raw_df, header_rows) in sheets.items():
            out, kinds, percent_cols, zscore_cols = _prepare_cross_section_sheet(raw_df)
            safe_name = sheet_name[:31]
            out.to_excel(writer, index=False, sheet_name=safe_name)
            _format_worksheet(
                writer.sheets[safe_name],
                out,
                kinds,
                as_table=True,
                table_name=_table_display_name(safe_name),
                full_scan=True,
                percent_cols=percent_cols,
                zscore_cols=zscore_cols,
                group_header_rows=header_rows,
            )
    return path


def write_excel(df: pd.DataFrame, path: str, sheet_name: str = "data") -> str:
    """Write ``df`` to ``path`` (.xlsx) with Excel-friendly date/time formatting
    and fitted column widths. Returns the path written."""
    out, kinds = _prepare_sheet(df)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name=sheet_name)
        _format_worksheet(writer.sheets[sheet_name], out, kinds)
    return path


def write_layers_workbook(
    path: str | Path,
    sheets: dict[str, pd.DataFrame],
    *,
    full_scan: bool = True,
) -> str:
    """Write narrative_layers workbook with presentation transforms and Excel tables."""
    path = str(path)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            out, kinds, percent_cols, zscore_cols = _prepare_layers_sheet(df, sheet_name)
            safe_name = sheet_name[:31]
            out.to_excel(writer, index=False, sheet_name=safe_name)
            _format_worksheet(
                writer.sheets[safe_name],
                out,
                kinds,
                as_table=True,
                table_name=_table_display_name(safe_name),
                full_scan=full_scan,
                percent_cols=percent_cols,
                zscore_cols=zscore_cols,
            )
    return path


def write_multi_sheet_excel(
    path: str | Path,
    sheets: dict[str, pd.DataFrame],
    *,
    as_tables: bool = False,
    full_scan: bool = False,
) -> str:
    """Write multiple DataFrames to one workbook (sheet names as keys)."""
    path = str(path)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            out, kinds = _prepare_sheet(df)
            safe_name = sheet_name[:31]
            out.to_excel(writer, index=False, sheet_name=safe_name)
            _format_worksheet(
                writer.sheets[safe_name],
                out,
                kinds,
                as_table=as_tables,
                table_name=_table_display_name(safe_name),
                full_scan=full_scan,
            )
    return path


NARRATIVE_LAYER_SHEETS = (
    ("narrative_zscored", "narrative_zscored"),
    ("dimension_scores", "dimension_scores"),
    ("dimension_surprise", "dimension_surprise"),
)


DEFAULT_DEMO_TICKERS = ("AMZN", "MSFT")
DEFAULT_DEMO_QUARTERS = ("FY2025-Q1", "FY2025-Q2")


def _load_layer(ticker: str, stem: str) -> pd.DataFrame:
    src = resolve_read_parquet_or_csv(ticker, stem, layer="parquet")
    if src is None:
        raise FileNotFoundError(
            f"Missing {stem} for {ticker.upper()}. "
            f"Run the pipeline for this ticker first."
        )
    return pd.read_parquet(src) if src.suffix == ".parquet" else pd.read_csv(src)


def _filter_quarters(df: pd.DataFrame, quarters: tuple[str, ...]) -> pd.DataFrame:
    if "fiscal_period" not in df.columns:
        return df
    allowed = set(quarters)
    return df[df["fiscal_period"].astype(str).isin(allowed)].copy()


def _load_narrative_layer_frames(
    ticker: str,
    quarters: tuple[str, ...],
) -> dict[str, pd.DataFrame]:
    t = ticker.upper()
    zscored = _filter_quarters(_load_layer(t, "narrative_zscored"), quarters)
    if "ticker" not in zscored.columns:
        zscored.insert(0, "ticker", t)

    dim_wide = _filter_quarters(_load_layer(t, "dimension_scores"), quarters)
    dim_scores = _dimension_scores_for_excel(dim_wide, t)

    surprise = _filter_quarters(_load_layer(t, "dimension_surprise"), quarters)
    if "ticker" not in surprise.columns:
        surprise.insert(0, "ticker", t)

    return {
        "narrative_zscored": zscored,
        "dimension_scores": dim_scores,
        "dimension_surprise": surprise,
    }


def build_consolidated_narrative_layers_workbook(
    tickers: list[str],
    quarters: tuple[str, ...] = DEFAULT_DEMO_QUARTERS,
    *,
    output_stem: str = "FY2025_Q1_Q2_narrative_layers",
) -> Path:
    """Stack narrative layers for multiple tickers, filter to selected quarters."""
    if not tickers:
        raise ValueError("At least one ticker is required.")

    stacked: dict[str, list[pd.DataFrame]] = {
        sheet: [] for _, sheet in NARRATIVE_LAYER_SHEETS
    }
    for ticker in tickers:
        frames = _load_narrative_layer_frames(ticker, quarters)
        for sheet, df in frames.items():
            if df.empty:
                print(f"  ! {ticker.upper()} {sheet}: no rows for {', '.join(quarters)}")
            stacked[sheet].append(df)

    out_frames = {
        sheet: pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
        for sheet, dfs in stacked.items()
    }
    for sheet, df in out_frames.items():
        sort_cols = [c for c in ("ticker", "fiscal_period", "dimension") if c in df.columns]
        if sort_cols:
            out_frames[sheet] = df.sort_values(sort_cols).reset_index(drop=True)

    dest = cross_company_artifact("workbooks", output_stem, "xlsx", mkdir=True)
    write_layers_workbook(dest, out_frames)
    return dest


def build_narrative_layers_workbook(ticker: str) -> Path:
    """Build output/{TICKER}/workbooks/narrative_layers.xlsx from parquet layers."""
    t = ticker.upper()
    frames: dict[str, pd.DataFrame] = {}
    for stem, sheet in NARRATIVE_LAYER_SHEETS:
        src = resolve_read_parquet_or_csv(t, stem, layer="parquet")
        if src is None:
            raise FileNotFoundError(
                f"Missing {stem}.parquet for {t}. "
                f"Run narrative_zscore / run_surprise_scoring first."
            )
        df = pd.read_parquet(src) if src.suffix == ".parquet" else pd.read_csv(src)
        if sheet == "dimension_scores":
            df = _dimension_scores_for_excel(df, t)
        elif "ticker" not in df.columns:
            df = df.copy()
            df.insert(0, "ticker", t)
        frames[sheet] = df

    dest = company_artifact(t, "workbooks", "narrative_layers", "xlsx", mkdir=True)
    write_layers_workbook(dest, frames)
    return dest


def _regen_layered_parquets(parquet_dir: Path, workbook_dir: Path) -> int:
    if not parquet_dir.is_dir():
        return 0
    workbook_dir.mkdir(parents=True, exist_ok=True)
    count = 0
    for pq in sorted(parquet_dir.glob("*.parquet")):
        df = pd.read_parquet(pq)
        xlsx = workbook_dir / f"{pq.stem}.xlsx"
        write_excel(df, str(xlsx))
        print(f"Wrote {xlsx}  ({len(df)} rows)")
        count += 1
    return count


def _regen_all(out_dir: str | Path | None = None):
    root = Path(out_dir) if out_dir is not None else ROOT
    total = 0
    for ticker in list_company_tickers():
        total += _regen_layered_parquets(
            root / ticker / "parquet",
            root / ticker / "workbooks",
        )
    total += _regen_layered_parquets(
        cross_company_layer("parquet"),
        cross_company_layer("workbooks", mkdir=True),
    )
    if total == 0:
        print(f"No layered .parquet files found under {root}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Excel export utilities.")
    ap.add_argument("--ticker", help="Ticker for --layers-workbook.")
    ap.add_argument(
        "--layers-workbook",
        action="store_true",
        help="Build workbooks/narrative_layers.xlsx for --ticker.",
    )
    ap.add_argument(
        "--consolidated-workbook",
        action="store_true",
        help="Build cross_company narrative_layers xlsx from existing parquet/csv.",
    )
    ap.add_argument(
        "--tickers",
        nargs="+",
        default=list(DEFAULT_DEMO_TICKERS),
        help="Tickers for --consolidated-workbook.",
    )
    ap.add_argument(
        "--quarters",
        nargs="+",
        default=list(DEFAULT_DEMO_QUARTERS),
        help="Fiscal periods to include (e.g. FY2025-Q1 FY2025-Q2).",
    )
    ap.add_argument(
        "--output-stem",
        default="FY2025_Q1_Q2_narrative_layers",
        help="Output filename stem under cross_company/.",
    )
    args = ap.parse_args()
    if args.consolidated_workbook:
        tickers = [t.upper() for t in args.tickers]
        quarters = tuple(q.upper() for q in args.quarters)
        path = build_consolidated_narrative_layers_workbook(
            tickers,
            quarters,
            output_stem=args.output_stem,
        )
        print(f"Wrote {path}  (tickers={', '.join(tickers)}, quarters={', '.join(quarters)})")
    elif args.layers_workbook:
        if not args.ticker:
            ap.error("--layers-workbook requires --ticker")
        path = build_narrative_layers_workbook(args.ticker)
        print(f"Wrote {path}")
    else:
        _regen_all()

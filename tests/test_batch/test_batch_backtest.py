import unittest
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from src.batch.models import BatchQuarterResult
from src.enrichment.enrichment_runner import run_quarter_enrichment
from src.enrichment.models import EnrichmentResult, TranscriptSource
from src.export.csv_writer import batch_result_to_excel_row, write_batch_excel
from src.ingest.documents.corpus_trim import trim_document_text
from src.ingest.documents.fetch.edgar_submissions import (
    FilingRecord,
    find_filings,
)
from src.ingest.documents.models import DocumentType, FetchedDocument
from src.market.quarter_labels import calendar_quarter_labels_back
from src.schemas.models import EvidenceClaim, LLMResult, QuarterSummary, TokenUsage


class BatchCutoffTestCase(unittest.TestCase):
    def test_find_filings_respects_knowledge_cutoff(self):
        filings = [
            FilingRecord(
                form="10-K",
                accession_number="a1",
                filing_date=date(2024, 3, 1),
                report_date=date(2023, 12, 31),
                primary_document="10k.htm",
                items=None,
            ),
            FilingRecord(
                form="10-K",
                accession_number="a2",
                filing_date=date(2025, 3, 1),
                report_date=date(2024, 12, 31),
                primary_document="10k.htm",
                items=None,
            ),
        ]
        matches = find_filings(
            filings,
            form="10-K",
            filed_on_or_before=date(2024, 8, 1),
        )
        self.assertEqual([record.accession_number for record in matches], ["a1"])


class BatchLabelTestCase(unittest.TestCase):
    def test_calendar_quarter_labels_back_40_quarters(self):
        labels = calendar_quarter_labels_back(
            40,
            end_label="2025-Q3",
        )
        self.assertEqual(len(labels), 40)
        self.assertEqual(labels[0], "2015-Q4")
        self.assertEqual(labels[-1], "2025-Q3")


class BatchPlaceholderExportTestCase(unittest.TestCase):
    def test_failed_batch_row_has_blank_scores(self):
        row = batch_result_to_excel_row(
            BatchQuarterResult(
                quarter_label="2018-Q1",
                status="failed",
                error_message="No earnings 8-K found",
                knowledge_cutoff=None,
            )
        )
        self.assertEqual(row["Confidence Score"], "")
        self.assertEqual(row["Document-Only Score"], "")
        self.assertIn("EDGAR fetch failed", row["What Happened"])


class BatchCorpusTrimTestCase(unittest.TestCase):
    def test_periodic_doc_is_capped(self):
        body = "Item 2. Management's Discussion\n" + ("detail " * 20_000)
        trimmed = trim_document_text(
            FetchedDocument(
                doc_type=DocumentType.TEN_Q,
                text=body,
            )
        )
        self.assertLessEqual(len(trimmed), 40_500)
        self.assertIn("Management's Discussion", trimmed)


class EnrichmentExportTestCase(unittest.TestCase):
    def test_batch_workbook_has_transcript_enrichment_sheet(self):
        summary = QuarterSummary(
            company_name="Amazon",
            quarter="2024-Q1",
            what_happened=["Revenue up"],
            positives=["Cloud growth"],
            negatives=["FX headwind"],
            transcript_only_confidence_score=10,
            confidence_score=10,
            analysis=[
                EvidenceClaim(
                    claim="+10: Revenue supports outlook",
                    excerpt="Revenue up.",
                )
            ],
        )
        results = [
            BatchQuarterResult(
                quarter_label="2024-Q1",
                status="success",
                summary=summary,
            )
        ]
        enrichment = [
            EnrichmentResult(
                quarter="2024-Q1",
                positives=[
                    EvidenceClaim(claim="AWS growth", excerpt="AWS grew 12%.")
                ],
                availability="found",
                notes="Source: local_cache",
            )
        ]
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "batch.xlsx"
            write_batch_excel(results, path, enrichment_results=enrichment)
            workbook = load_workbook(path)
            self.assertIn("Transcript Enrichment", workbook.sheetnames)
            self.assertIn(
                "Edgar documents only",
                str(workbook["Batch Backtest"].cell(1, 1).value),
            )


class EnrichmentIsolationTestCase(unittest.TestCase):
    def test_enrichment_runner_uses_separate_prompt(self):
        client = MagicMock()
        client.complete_json.return_value = (
            MagicMock(
                positives=[EvidenceClaim(claim="AWS growth", excerpt="AWS grew 12%.")],
                negatives=[],
                key_quotes=[],
                availability="found",
            ),
            LLMResult(usage=TokenUsage(input_tokens=1, output_tokens=1), raw_response="{}"),
        )
        with patch(
            "src.enrichment.enrichment_runner.fetch_transcript",
            return_value=TranscriptSource(
                quarter="2024-Q1",
                text="AWS grew 12%.",
                source="local_cache",
            ),
        ):
            result = run_quarter_enrichment(
                client,
                ticker="AMZN",
                quarter_label="2024-Q1",
            )
        self.assertEqual(result.availability, "found")
        system_prompt = client.complete_json.call_args.kwargs["system_prompt"]
        self.assertIn("informational", system_prompt.lower())


if __name__ == "__main__":
    unittest.main()

import unittest
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from src.batch.models import BatchQuarterResult
from src.enrichment.models import EnrichmentResult
from src.export.csv_writer import write_batch_excel
from src.schemas.models import EvidenceBackedQuarterSummary, EvidenceClaim, QuarterSummary


class ValidatedFactorsExportTestCase(unittest.TestCase):
    def test_validated_factors_sheet_merges_lanes(self):
        filing_evidence = EvidenceBackedQuarterSummary(
            company_name="Nvidia",
            quarter="FY2025-Q2",
            what_happened=[
                EvidenceClaim(claim="Record revenue", excerpt="record revenue for the quarter.")
            ],
            positives=[
                EvidenceClaim(claim="Data center strength", excerpt="Data center revenue grew.")
            ],
            negatives=[
                EvidenceClaim(claim="Supply constraints", excerpt="supply remained constrained.")
            ],
            analysis=[
                EvidenceClaim(
                    claim="+20: Data center demand",
                    excerpt="Data center revenue grew.",
                )
            ],
            transcript_only_confidence_score=20,
            confidence_score=20,
        )
        summary = QuarterSummary(
            company_name="Nvidia",
            quarter="FY2025-Q2",
            what_happened=["Record revenue"],
            positives=["Data center strength"],
            negatives=["Supply constraints"],
            transcript_only_confidence_score=20,
            confidence_score=20,
            analysis=filing_evidence.analysis,
        )
        enrichment = EnrichmentResult(
            quarter="FY2025-Q2",
            positives=[
                EvidenceClaim(claim="Call tone bullish", excerpt="demand remains strong.")
            ],
            negatives=[],
            availability="found",
            notes="Source: local cache",
            validation_status="kept=1",
        )
        batch_result = BatchQuarterResult(
            quarter_label="FY2025-Q2",
            status="success",
            summary=summary,
            filing_evidence=filing_evidence,
            enrichment=enrichment,
            fetch_summary="PR + Commentary + 8-K; Transcript(found)",
        )

        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "batch.xlsx"
            write_batch_excel([batch_result], path, enrichment_results=[enrichment])
            workbook = load_workbook(path)
            self.assertIn("Validated Factors", workbook.sheetnames)
            sheet = workbook["Validated Factors"]
            header_row = 2
            rows = [
                [sheet.cell(row_index, col).value for col in range(1, 7)]
                for row_index in range(header_row + 1, sheet.max_row + 1)
            ]
            sources = {row[1] for row in rows}
            in_score_values = {row[5] for row in rows}
            self.assertIn("Filing", sources)
            self.assertIn("Transcript", sources)
            self.assertIn("Yes", in_score_values)
            self.assertIn("No", in_score_values)
            analysis_rows = [row for row in rows if row[2] == "analysis" and row[1] == "Filing"]
            self.assertTrue(analysis_rows)
            self.assertEqual(analysis_rows[0][5], "Yes")


if __name__ == "__main__":
    unittest.main()

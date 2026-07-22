"""Tests for slim spine export and consolidated panel methodology."""
from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SN = ROOT / "Structured Narrative"
sys.path.insert(0, str(SN))

from excel_export import write_cross_section_panel_workbook  # noqa: E402
from panel_html import EvidenceLookups, build_consolidated_html, summarize_ticker_quarter  # noqa: E402
from period_dates import (  # noqa: E402
    apply_feature_availability_dates,
    apply_investable_cross_section_columns,
    calendar_quarter_from_date,
    enrich_panel_period_columns,
    filter_min_calendar_quarter,
    model_date_from,
    quarter_cell_html,
)
from spine_export import (  # noqa: E402
    CONSOLIDATED_SPINE_COLUMNS,
    panel_to_spine,
    standardize_surprise_novelty_exclusivity,
    validate_spine_rules,
)
from quant_mapping import quant_mapping_for, quant_family_for  # noqa: E402


def _sample_panel(
    ticker: str,
    fp: str,
    *,
    period_end: str = "2025-03-31",
    earnings_date: str = "2025-05-01",
) -> pd.DataFrame:
    bucket = calendar_quarter_from_date(period_end)
    rows = [
        {
            "ticker": ticker,
            "fiscal_period": fp,
            "period_end_date": period_end,
            "period_end_calendar_quarter": bucket,
            "dimension": "demand",
            "earnings_date": earnings_date,
            "call_feature_available_date": earnings_date,
            "t7_feature_available_date": None,
            "feature_availability_date": earnings_date,
            "quant_mapping": quant_mapping_for("demand"),
            "quant_family": quant_family_for("demand"),
            "llm_level": 1.5,
            "change_direction": "improved",
            "change_magnitude": 0.2,
            "surprise_magnitude": 0.8,
            "narrative_novelty": None,
            "quant_z_pit": -0.4,
            "quant_z": -0.4,
            "agrees_with_quant": False,
            "narrative_quant_gap": 1.2,
            "level_evidence_supported_pct": 1.0,
            "delta_evidence_supported_pct": 1.0,
            "surprise_evidence_supported_pct": 0.9,
            "is_divergence": True,
            "any_quant_divergence": True,
            "has_delta": True,
            "has_level": True,
            "has_surprise": True,
            "has_novelty": False,
            "signal_stack": "bullish_level|surprise_bullish|quant_diverges",
            "level_rationale": "Strong demand cited.",
            "delta_rationale": "Improved sequentially.",
            "surprise_rationale": "Beat expectations.",
        },
        {
            "ticker": ticker,
            "fiscal_period": fp,
            "period_end_date": period_end,
            "period_end_calendar_quarter": bucket,
            "dimension": "management_confidence",
            "earnings_date": earnings_date,
            "call_feature_available_date": earnings_date,
            "t7_feature_available_date": None,
            "feature_availability_date": earnings_date,
            "quant_mapping": "",
            "quant_family": None,
            "llm_level": 0.5,
            "change_magnitude": 0.0,
            "surprise_magnitude": None,
            "narrative_novelty": 1.2,
            "novelty_direction": "high_novelty",
            "quant_z_pit": None,
            "agrees_with_quant": None,
            "level_evidence_supported_pct": 1.0,
            "delta_evidence_supported_pct": 1.0,
            "novelty_evidence_supported_pct": 1.0,
            "is_divergence": False,
            "has_novelty": True,
            "has_level": True,
            "has_delta": True,
            "level_rationale": "Confident tone.",
            "novelty_rationale": "New AI disclosure.",
        },
    ]
    return pd.DataFrame(rows)


class SpineExportTests(unittest.TestCase):
    def test_panel_to_spine_columns(self):
        spine = panel_to_spine(_sample_panel("TST", "FY2025-Q1"))
        for col in CONSOLIDATED_SPINE_COLUMNS:
            self.assertIn(col, spine.columns)

    def test_surprise_novelty_mutual_exclusivity(self):
        spine = panel_to_spine(_sample_panel("TST", "FY2025-Q1"))
        errors = validate_spine_rules(spine)
        self.assertEqual(errors, [])

    def test_quant_mapping_nonempty_for_demand(self):
        self.assertIn("Sales", quant_mapping_for("demand"))


class PeriodEndBucketTests(unittest.TestCase):
    def test_amzn_msft_fy2025_q1_separate_buckets(self):
        amzn = _sample_panel("AMZN", "FY2025-Q1", period_end="2025-03-31")
        msft = _sample_panel("MSFT", "FY2025-Q1", period_end="2024-09-30")
        self.assertEqual(amzn.iloc[0]["period_end_calendar_quarter"], "2025-Q1")
        self.assertEqual(msft.iloc[0]["period_end_calendar_quarter"], "2024-Q3")

    def test_exact_calendar_quarter_ends_unchanged(self):
        self.assertEqual(calendar_quarter_from_date("2021-09-30"), "2021-Q3")
        self.assertEqual(calendar_quarter_from_date("2021-12-31"), "2021-Q4")
        self.assertEqual(calendar_quarter_from_date("2022-03-31"), "2022-Q1")

    def test_nvda_late_month_ends_nearest_prior_quarter(self):
        # NVIDIA FY ends ~Jan 31; Q3 ends ~Oct 31 — bucket with Sep/Dec peers.
        self.assertEqual(calendar_quarter_from_date("2021-10-31"), "2021-Q3")
        self.assertEqual(calendar_quarter_from_date("2022-01-31"), "2021-Q4")
        self.assertEqual(calendar_quarter_from_date("2022-04-30"), "2022-Q1")
        self.assertEqual(calendar_quarter_from_date("2022-07-31"), "2022-Q2")

    def test_earnings_date_calendar_quarter_bucketed_from_earnings_date(self):
        # period_end 2025-03-31 (-> 2025-Q1) but earnings_date 2025-05-01 is nearest
        # to 2025-03-31 too (31d away vs. 60d to 2025-06-30) — still Q1 here; the
        # point is the column is derived from earnings_date, not period_end_date.
        panel = _sample_panel("AMZN", "FY2025-Q1", period_end="2025-03-31", earnings_date="2025-05-01")
        enriched = enrich_panel_period_columns(panel)
        self.assertIn("earnings_date_calendar_quarter", enriched.columns)
        self.assertEqual(enriched.iloc[0]["earnings_date_calendar_quarter"], "2025-Q1")

    def test_earnings_date_calendar_quarter_diverges_from_period_end_bucket(self):
        # period ends 2025-03-31 (-> Q1) but the call itself happens late enough
        # (2025-06-20) to land nearer the following calendar quarter-end.
        panel = _sample_panel("AMZN", "FY2025-Q1", period_end="2025-03-31", earnings_date="2025-06-20")
        enriched = enrich_panel_period_columns(panel)
        self.assertEqual(enriched.iloc[0]["period_end_calendar_quarter"], "2025-Q1")
        self.assertEqual(enriched.iloc[0]["earnings_date_calendar_quarter"], "2025-Q2")

    def test_model_date_from_adds_seven_calendar_days(self):
        self.assertEqual(model_date_from("2025-05-01").isoformat(), "2025-05-08")

    def test_model_date_from_rolls_off_weekend(self):
        # 2025-04-26 is a Saturday; +7d lands on another Saturday (2025-05-03),
        # which rolls forward to the following Monday.
        self.assertEqual(model_date_from("2025-04-26").isoformat(), "2025-05-05")

    def test_filter_min_calendar_quarter_drops_early_amzn(self):
        early = _sample_panel("AMZN", "FY2021-Q2", period_end="2021-06-30")
        keep = _sample_panel("AMZN", "FY2021-Q3", period_end="2021-09-30")
        stacked = pd.concat([early, keep], ignore_index=True)
        stacked = enrich_panel_period_columns(stacked)
        trimmed = filter_min_calendar_quarter(stacked, "2021-Q3")
        self.assertEqual(set(trimmed["fiscal_period"]), {"FY2021-Q3"})

    def test_quarter_cell_html_period_ending_subline(self):
        row = _sample_panel("AMZN", "FY2025-Q1").iloc[0]
        html = quarter_cell_html(row)
        self.assertIn("Period ending", html)
        self.assertIn("03/31/2025", html)
        self.assertIn("Earnings call", html)

    def test_guidance_feature_availability_feature_level(self):
        row = {
            "dimension": "guidance",
            "earnings_date": "2025-05-01",
            "as_of_date": "2025-05-01",
            "model_date": "2025-05-08",
            "quant_guidance_revision_z_pit": 0.5,
        }
        panel = apply_feature_availability_dates(pd.DataFrame([row]))
        r = panel.iloc[0]
        self.assertEqual(str(r["call_feature_available_date"])[:10], "2025-05-01")
        self.assertEqual(str(r["t7_feature_available_date"])[:10], "2025-05-08")
        self.assertEqual(str(r["feature_availability_date"])[:10], "2025-05-01")

    def test_demand_has_no_t7_availability(self):
        row = {
            "dimension": "demand",
            "earnings_date": "2025-05-01",
            "as_of_date": "2025-05-01",
            "quant_guidance_revision_z_pit": None,
        }
        panel = apply_feature_availability_dates(pd.DataFrame([row]))
        self.assertEqual(str(panel.iloc[0]["call_feature_available_date"])[:10], "2025-05-01")
        self.assertTrue(pd.isna(panel.iloc[0]["t7_feature_available_date"]))

    def test_investable_as_of_uses_latest_earnings_t7(self):
        stacked = pd.concat(
            [
                _sample_panel("AAA", "FY2025-Q1", period_end="2025-03-31", earnings_date="2025-04-20"),
                _sample_panel("BBB", "FY2025-Q1", period_end="2025-03-31", earnings_date="2025-05-01"),
            ],
            ignore_index=True,
        )
        out = apply_investable_cross_section_columns(stacked)
        # Latest earnings 2025-05-01 → as-of 2025-05-08
        self.assertTrue((out["investable_as_of_date"].astype(str).str[:10] == "2025-05-08").all())
        aaa = out[out["ticker"] == "AAA"].iloc[0]
        self.assertEqual(int(aaa["days_since_earnings"]), 18)
        self.assertTrue(bool(aaa["investable_ready"]))

    def test_legacy_surprise_novelty_cleared(self):
        dirty = _sample_panel("AMZN", "FY2025-Q1")
        dirty.loc[dirty["dimension"] == "management_confidence", "surprise_magnitude"] = 0.5
        dirty.loc[dirty["dimension"] == "demand", "narrative_novelty"] = 1.0
        clean = standardize_surprise_novelty_exclusivity(dirty)
        narr = clean[clean["dimension"] == "management_confidence"].iloc[0]
        dem = clean[clean["dimension"] == "demand"].iloc[0]
        self.assertTrue(pd.isna(narr["surprise_magnitude"]))
        self.assertTrue(pd.isna(dem["narrative_novelty"]))
        self.assertEqual(validate_spine_rules(panel_to_spine(clean)), [])

    def test_compare_filter_uses_period_bucket(self):
        stacked = pd.concat(
            [
                _sample_panel("AMZN", "FY2025-Q1", period_end="2025-03-31"),
                _sample_panel("MSFT", "FY2025-Q1", period_end="2024-09-30"),
            ],
            ignore_index=True,
        )
        empty = EvidenceLookups(level={}, delta={}, surprise={})
        html = build_consolidated_html(
            stacked,
            {"AMZN": empty, "MSFT": empty},
            tickers=["AMZN", "MSFT"],
            period_buckets=["2025-Q1", "2024-Q3"],
            default_bucket="2025-Q1",
            sector_label=None,
            generated_at="2026-01-01T00:00:00Z",
        )
        self.assertIn('data-period-bucket="2025-Q1"', html)
        self.assertIn('data-period-bucket="2024-Q3"', html)
        self.assertIn("calendar quarter of fiscal period-end", html)
        self.assertIn("Period ending", html)


class ConsolidatedPanelReportTests(unittest.TestCase):
    def test_summarize_ticker_quarter(self):
        stats = summarize_ticker_quarter(_sample_panel("TST", "FY2025-Q1"))
        self.assertEqual(stats["divergence_count"], 1)
        self.assertIsNotNone(stats["level_avg"])

    def test_build_consolidated_html_markers(self):
        from dimension_order import prepare_consolidated_panel

        stacked = prepare_consolidated_panel(
            pd.concat(
                [_sample_panel("AAA", "FY2025-Q1"), _sample_panel("BBB", "FY2025-Q1")],
                ignore_index=True,
            )
        )
        empty = EvidenceLookups(level={}, delta={}, surprise={})
        html = build_consolidated_html(
            stacked,
            {"AAA": empty, "BBB": empty},
            tickers=["AAA", "BBB"],
            period_buckets=["2025-Q1"],
            default_bucket="2025-Q1",
            sector_label="test_sector",
            generated_at="2026-01-01T00:00:00Z",
        )
        self.assertIn('data-ticker="AAA"', html)
        self.assertIn("dim-group-header", html)
        self.assertIn("Fundamentals", html)
        self.assertIn("Quant PIT", html)
        self.assertIn("Novelty", html)
        self.assertIn('data-mode="compare"', html)
        self.assertIn("Call features", html)
        self.assertIn("Investable as-of", html)

    def test_build_script_runs_if_panels_exist(self):
        """Smoke-test the CLI entry point with a narrow --tickers MSFT run.

        This MUST NOT write into the shared output/cross_company/{reports,json,
        csv,workbooks}/ tree: that is where the real 4-ticker
        consolidated_feature_panel.* deliverable lives. A prior version of this
        test ran the script against the production output paths, which meant
        every full-suite run silently overwrote the real consolidated report
        with MSFT-only data (the recurring "report reverted to MSFT-only" bug).
        Redirect cross-company output to an isolated temp dir via
        SN_CROSS_COMPANY_OUTPUT_DIR (see output_paths.py) so this test can never
        clobber the real deliverable, no matter how many times it runs.
        """
        script = SN / "build_consolidated_panel_report.py"
        msft_panel = SN / "output" / "MSFT" / "csv" / "feature_panel.csv"
        if not msft_panel.exists():
            self.skipTest("MSFT feature panel not present")
        import os
        import subprocess

        with tempfile.TemporaryDirectory() as tmp_cross:
            env = dict(os.environ)
            env["SN_CROSS_COMPANY_OUTPUT_DIR"] = tmp_cross
            result = subprocess.run(
                [sys.executable, str(script), "--tickers", "MSFT"],
                cwd=ROOT,
                capture_output=True,
                text=True,
                env=env,
            )
            self.assertEqual(result.returncode, 0, msg=result.stderr)
            # Confirm the isolation actually took effect (wrote to the temp
            # dir, not to the real output/cross_company/ tree).
            self.assertTrue(
                (Path(tmp_cross) / "reports" / "consolidated_feature_panel.html").exists()
            )


class CrossSectionPanelWorkbookTests(unittest.TestCase):
    def test_write_cross_section_panel_workbook_tables_and_rows(self):
        panel = pd.concat(
            [
                _sample_panel("AAA", "FY2025-Q1", period_end="2025-03-31"),
                _sample_panel("BBB", "FY2025-Q1", period_end="2025-03-31"),
            ],
            ignore_index=True,
        )
        from dimension_order import prepare_consolidated_panel

        panel = prepare_consolidated_panel(panel)
        spine = panel_to_spine(panel)
        expected_rows = len(panel)

        with tempfile.TemporaryDirectory() as tmp:
            xlsx_path = Path(tmp) / "cross_section_panel.xlsx"
            write_cross_section_panel_workbook(xlsx_path, panel, spine)

            wb = load_workbook(xlsx_path)
            self.assertEqual(set(wb.sheetnames), {"Summary", "Panel"})

            for sheet_name in ("Summary", "Panel"):
                ws = wb[sheet_name]
                self.assertEqual(ws.freeze_panes, "A2")
                # Summary uses autofilter (group headers); Panel uses Excel table or autofilter.
                has_filter = ws.auto_filter.ref is not None or len(ws.tables) >= 1
                self.assertTrue(has_filter, msg=sheet_name)

            # Panel is observation-only (no Fundamentals / Narrative Context header rows).
            self.assertEqual(wb["Panel"].max_row, expected_rows + 1)
            panel_a_vals = {
                str(wb["Panel"]["A" + str(r)].value)
                for r in range(2, wb["Panel"].max_row + 1)
                if wb["Panel"]["A" + str(r)].value
            }
            self.assertNotIn("Fundamentals", panel_a_vals)
            self.assertNotIn("Narrative context", panel_a_vals)

            # Summary keeps formatted group headers.
            self.assertGreater(wb["Summary"].max_row, expected_rows + 1)
            summary_labels = {
                str(wb["Summary"]["A" + str(r)].value)
                for r in range(2, wb["Summary"].max_row + 1)
                if wb["Summary"]["A" + str(r)].value
            }
            self.assertIn("Fundamentals", summary_labels)
            self.assertIn("Narrative context", summary_labels)


if __name__ == "__main__":
    unittest.main()

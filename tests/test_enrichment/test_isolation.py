import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from src.batch.models import BatchQuarterResult
from src.enrichment.enrichment_runner import run_quarter_enrichment
from src.enrichment.models import EnrichmentResult, TranscriptSource
from src.export.csv_writer import batch_result_to_excel_row, write_batch_excel
from src.schemas.models import EvidenceClaim, LLMResult, QuarterSummary, TokenUsage


class EnrichmentIsolationTestCase(unittest.TestCase):
    def test_enrichment_does_not_change_confidence_row(self):
        summary = QuarterSummary(
            company_name="Amazon",
            quarter="2024-Q1",
            what_happened=["Revenue up"],
            positives=["Cloud growth"],
            negatives=["FX headwind"],
            transcript_only_confidence_score=12,
            confidence_score=12,
            analysis=[
                EvidenceClaim(
                    claim="+12: Revenue supports outlook",
                    excerpt="Revenue up.",
                )
            ],
        )
        batch_result = BatchQuarterResult(
            quarter_label="2024-Q1",
            status="success",
            summary=summary,
        )
        confidence_row = batch_result_to_excel_row(batch_result)

        enrichment = EnrichmentResult(
            quarter="2024-Q1",
            positives=[EvidenceClaim(claim="Different lane", excerpt="AWS grew 40%.")],
            negatives=[EvidenceClaim(claim="Separate lane", excerpt="Costs rose.")],
            availability="found",
            notes="Source: local cache",
            validation_status="kept=2",
        )
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "batch.xlsx"
            write_batch_excel([batch_result], path, enrichment_results=[enrichment])
            workbook = load_workbook(path)
            backtest = workbook["Batch Backtest"]
            header_row = 2 if "Edgar" in str(backtest.cell(1, 1).value) else 1
            data_row = header_row + 1
            headers = [backtest.cell(header_row, col).value for col in range(1, 12)]
            col = {header: index + 1 for index, header in enumerate(headers) if header}
            self.assertEqual(
                backtest.cell(data_row, col["Confidence Score"]).value,
                confidence_row["Confidence Score"],
            )
            self.assertEqual(
                backtest.cell(data_row, col["Positives"]).value,
                confidence_row["Positives"],
            )
            self.assertNotIn(
                "AWS grew 40%",
                str(backtest.cell(data_row, col["Positives"]).value or ""),
            )

    def test_enrichment_runner_uses_separate_prompt(self):
        client = MagicMock()
        client.complete_json.return_value = (
            MagicMock(
                positives=[EvidenceClaim(claim="AWS growth", excerpt="AWS grew 12%.")],
                negatives=[],
                key_quotes=[],
                availability="found",
            ),
            LLMResult(usage=TokenUsage(input_tokens=1, output_tokens=1), raw_response="{}"),
        )
        with patch(
            "src.enrichment.enrichment_runner.fetch_transcript",
            return_value=TranscriptSource(
                quarter="2024-Q1",
                text="AWS grew 12%.",
                source="local cache",
            ),
        ):
            result = run_quarter_enrichment(
                client,
                ticker="AMZN",
                quarter_label="2024-Q1",
            )
        self.assertEqual(result.availability, "found")
        system_prompt = client.complete_json.call_args.kwargs["system_prompt"]
        self.assertIn("informational", system_prompt.lower())
        self.assertIn(
            "never be used to compute or influence a confidence score",
            system_prompt.lower(),
        )

    def test_enrichment_rejects_non_verbatim_excerpts(self):
        client = MagicMock()
        client.complete_json.return_value = (
            MagicMock(
                positives=[EvidenceClaim(claim="AWS growth", excerpt="AWS grew 99%.")],
                negatives=[],
                key_quotes=[],
                availability="found",
            ),
            LLMResult(usage=TokenUsage(input_tokens=1, output_tokens=1), raw_response="{}"),
        )
        with patch(
            "src.enrichment.enrichment_runner.fetch_transcript",
            return_value=TranscriptSource(
                quarter="2024-Q1",
                text="AWS grew 12%.",
                source="local cache",
            ),
        ):
            result = run_quarter_enrichment(
                client,
                ticker="AMZN",
                quarter_label="2024-Q1",
            )
        self.assertEqual(result.positives, [])
        self.assertIn("dropped", result.validation_status or "")


if __name__ == "__main__":
    unittest.main()

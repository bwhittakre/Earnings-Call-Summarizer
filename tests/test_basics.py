import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from src.export.confidence_reference_key import (
    REFERENCE_KEY_TITLE,
    load_reference_key_text,
)
from src.export.csv_writer import (
    EXCEL_COLUMN_WIDTHS,
    EXCEL_MAX_ROW_HEIGHT,
    format_analysis_bullets,
    format_bullets,
    format_list,
    format_quarter_cell,
    format_what_happened,
    max_lines_at_max_height,
    min_column_width_for_text,
    summary_to_excel_row,
    summary_to_row,
    write_excel,
)
from src.schemas.models import EvidenceClaim, QuarterSummary

SAMPLE_ANALYSIS = [
    EvidenceClaim(
        claim="+20: Raised outlook supports next-quarter momentum",
        excerpt="We are raising full-year revenue guidance across every segment.",
    )
]


def _sample_summary(**overrides) -> QuarterSummary:
    payload = {
        "company_name": "Nvidia",
        "quarter": "FY2025-Q2",
        "what_happened": ["Strong data center demand", "Raised guidance"],
        "positives": ["Blackwell ramp", "Margin recovery"],
        "negatives": ["China restrictions"],
        "confidence_score": 72,
        "document_only_confidence_score": 72,
        "analysis": list(SAMPLE_ANALYSIS),
    }
    payload.update(overrides)
    return QuarterSummary(**payload)


class BasicsTestCase(unittest.TestCase):
    def test_csv_row_formatting(self):
        summary = _sample_summary(company_name="Microsoft", quarter="2024-Q3")
        row = summary_to_row(summary)
        self.assertEqual(row["what_happened"], "Strong data center demand & Raised guidance")
        self.assertEqual(row["positives"], "Blackwell ramp, Margin recovery")
        self.assertEqual(row["confidence_score"], "72")
        self.assertIn("We are raising full-year revenue guidance", row["analysis"])
        self.assertEqual(format_what_happened(["A", "B"]), "A & B")
        self.assertEqual(format_list(["X", "Y"]), "X, Y")
        self.assertEqual(format_bullets(["X", "Y"]), "- X\n- Y")

    def test_analysis_bullet_formatting_includes_excerpt(self):
        formatted = format_analysis_bullets(SAMPLE_ANALYSIS)
        self.assertIn("+20: Raised outlook supports next-quarter momentum", formatted)
        self.assertIn('"We are raising full-year revenue guidance across every segment."', formatted)
        self.assertIn(" — ", formatted)

    def test_min_column_width_for_text_expands_for_long_content(self):
        short_text = "Brief analysis point."
        long_text = "\n".join(
            f'• +10: Factor {index} — "{("detail " * 30).strip()}"'
            for index in range(15)
        )
        max_lines = max_lines_at_max_height()
        short_width = min_column_width_for_text(short_text, max_lines)
        long_width = min_column_width_for_text(long_text, max_lines)
        self.assertLess(short_width, long_width)
        self.assertGreater(long_width, EXCEL_COLUMN_WIDTHS["Analysis"])

    def test_write_excel_widens_analysis_column_for_long_analysis(self):
        long_analysis = [
            EvidenceClaim(
                claim=f"+{index}: Industry driver {index}",
                excerpt=("Management discussed metric with detailed commentary " + ("data " * 40)).strip(),
            )
            for index in range(1, 16)
        ]
        summary = _sample_summary(analysis=long_analysis)
        with TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "summary.xlsx"
            write_excel([summary], output_path)
            workbook = load_workbook(output_path)
            worksheet = workbook["Nvidia"]
            self.assertGreater(
                worksheet.column_dimensions["I"].width,
                EXCEL_COLUMN_WIDTHS["Analysis"],
            )
            self.assertLessEqual(worksheet.row_dimensions[2].height, EXCEL_MAX_ROW_HEIGHT)

    def test_excel_row_formatting(self):
        summary = _sample_summary()
        row = summary_to_excel_row(summary)
        self.assertEqual(row["Summary Type"], "Quarter")
        self.assertEqual(row["What Happened"], "- Strong data center demand\n- Raised guidance")
        self.assertEqual(row["Confidence Score"], "72")
        self.assertEqual(row["Document-Only Score"], "72")
        self.assertIn('"We are raising full-year revenue guidance across every segment."', row["Analysis"])

    def test_write_excel(self):
        summary = _sample_summary(as_of_date="(08,28,2024)")
        with TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "summary.xlsx"
            write_excel([summary], output_path)
            workbook = load_workbook(output_path)
            worksheet = workbook["Nvidia"]
            self.assertEqual(worksheet["A1"].value, "Summary Type")
            self.assertEqual(worksheet["G1"].value, "Document-Only Score")
            self.assertEqual(worksheet["H1"].value, "Confidence Score")
            self.assertEqual(worksheet["I1"].value, "Analysis")
            self.assertEqual(worksheet["D2"].value, "- Strong data center demand\n- Raised guidance")
            self.assertEqual(worksheet["G2"].value, "72")
            self.assertEqual(worksheet["H2"].value, "72")
            self.assertEqual(
                worksheet["C2"].value,
                "FY2025-Q2\nAs-of Date: (08,28,2024)",
            )
            self.assertEqual(worksheet["K1"].value, REFERENCE_KEY_TITLE)
            self.assertIn(
                "sum of all Analysis bullet weights",
                str(worksheet["K2"].value),
            )
            self.assertEqual(worksheet.auto_filter.ref, "A1:I2")

    def test_load_reference_key_text(self):
        text = load_reference_key_text()
        self.assertIn("Score interpretation bands", text)
        self.assertIn("no fixed maximum", text)
        self.assertIn("Strong stock-moving drivers: ±20 to ±25", text)
        self.assertIn("NEXT QUARTER ONLY", text)
        self.assertIn("four fiscal quarters immediately before", text)
        self.assertGreater(len(text), 100)

    def test_write_excel_creates_one_sheet_per_company(self):
        nvidia = _sample_summary()
        amazon = _sample_summary(
            company_name="Amazon",
            what_happened=["AWS growth improved"],
            positives=["Retail margin expansion"],
            negatives=["Capex pressure"],
            confidence_score=-15,
        )
        with TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "summary.xlsx"
            write_excel([nvidia, amazon], output_path)
            workbook = load_workbook(output_path)
            self.assertEqual(workbook.sheetnames, ["Amazon", "Nvidia"])
            self.assertEqual(workbook["Nvidia"]["B2"].value, "Nvidia")
            self.assertEqual(workbook["Amazon"]["B2"].value, "Amazon")
            self.assertEqual(workbook["Nvidia"]["K1"].value, REFERENCE_KEY_TITLE)
            self.assertEqual(workbook["Amazon"]["K1"].value, REFERENCE_KEY_TITLE)

    def test_write_excel_consolidates_multiple_quarters_on_one_sheet(self):
        rows = [
            _sample_summary(quarter="FY2026-Q1", confidence_score=42),
            _sample_summary(quarter="FY2026-Q2", confidence_score=46),
            _sample_summary(quarter="FY2026-Q3", confidence_score=67),
            _sample_summary(quarter="FY2026-Q4", confidence_score=90),
        ]
        with TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "summary.xlsx"
            write_excel(rows, output_path)
            workbook = load_workbook(output_path)
            self.assertEqual(workbook.sheetnames, ["Nvidia"])
            worksheet = workbook["Nvidia"]
            self.assertEqual(worksheet["C2"].value, "FY2026-Q1")
            self.assertEqual(worksheet["C3"].value, "FY2026-Q2")
            self.assertEqual(worksheet["C4"].value, "FY2026-Q3")
            self.assertEqual(worksheet["C5"].value, "FY2026-Q4")
            self.assertEqual(worksheet["H5"].value, "90")
            self.assertEqual(worksheet.auto_filter.ref, "A1:I5")

    def test_write_excel_single_sheet_combines_companies(self):
        rows = [
            _sample_summary(company_name="Nvidia", quarter="FY2026-Q2"),
            _sample_summary(
                company_name="Amazon",
                quarter="FY2026-Q2",
                confidence_score=10,
            ),
            _sample_summary(
                company_name="Tesla",
                quarter="FY2026-Q2",
                confidence_score=20,
            ),
        ]
        with TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "summary.xlsx"
            write_excel(rows, output_path, single_sheet=True)
            workbook = load_workbook(output_path)
            self.assertEqual(workbook.sheetnames, ["Earnings Summary"])
            worksheet = workbook["Earnings Summary"]
            self.assertEqual(worksheet["B2"].value, "Amazon")
            self.assertEqual(worksheet["B3"].value, "Nvidia")
            self.assertEqual(worksheet["B4"].value, "Tesla")
            self.assertEqual(worksheet.auto_filter.ref, "A1:I4")

    def test_format_quarter_cell(self):
        self.assertEqual(format_quarter_cell("FY2025-Q2"), "FY2025-Q2")
        self.assertEqual(
            format_quarter_cell("FY2025-Q2", "(08,28,2024)"),
            "FY2025-Q2\nAs-of Date: (08,28,2024)",
        )


if __name__ == "__main__":
    unittest.main()

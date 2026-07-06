from __future__ import annotations

import csv
import math
import re
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.export.confidence_reference_key import write_confidence_reference_key
from src.ingest.filings.fiscal import quarter_sort_key
from src.schemas.models import EvidenceClaim, QuarterSummary

CSV_COLUMNS = [
    "summary_type",
    "company_name",
    "quarter",
    "what_happened",
    "positives",
    "negatives",
    "document_only_confidence_score",
    "confidence_score",
    "analysis",
]

DISPLAY_HEADERS = {
    "summary_type": "Summary Type",
    "company_name": "Company Name",
    "quarter": "Quarter",
    "what_happened": "What Happened",
    "positives": "Positives",
    "negatives": "Negatives",
    "document_only_confidence_score": "Document-Only Score",
    "confidence_score": "Confidence Score",
    "analysis": "Analysis",
}

EXCEL_COLUMN_WIDTHS = {
    "Summary Type": 16,
    "Company Name": 18,
    "Quarter": 18,
    "What Happened": 42,
    "Positives": 60,
    "Negatives": 60,
    "Document-Only Score": 18,
    "Confidence Score": 16,
    "Analysis": 80,
}

EXCEL_MAX_ROW_HEIGHT = 409.6
EXCEL_MAX_COLUMN_WIDTH = 255
MIN_DATA_ROW_HEIGHT = 42
POINTS_PER_WRAPPED_LINE = 15
ROW_HEIGHT_PADDING = 8

TABLE_COLUMN_COUNT = len(CSV_COLUMNS)
REFERENCE_KEY_SPACER_COLUMN = TABLE_COLUMN_COUNT + 1
REFERENCE_KEY_START_COLUMN = TABLE_COLUMN_COUNT + 2


def format_quarter_cell(quarter: str, as_of_date: str | None = None) -> str:
    if as_of_date:
        return f"{quarter}\nAs-of Date: {as_of_date}"
    return quarter


def format_what_happened(items: list[str]) -> str:
    return " & ".join(items)


def format_list(items: list[str]) -> str:
    return ", ".join(items)


def format_bullets(items: list[str]) -> str:
    return "\n".join(f"• {item}" for item in items)


def format_analysis_bullets(items: list[EvidenceClaim]) -> str:
    return "\n".join(f'• {item.claim} — "{item.excerpt}"' for item in items)


def format_analysis_csv(items: list[EvidenceClaim]) -> str:
    return " | ".join(f'{item.claim} — "{item.excerpt}"' for item in items)


def summary_to_row(summary: QuarterSummary) -> dict[str, str]:
    return {
        "summary_type": summary.summary_type,
        "company_name": summary.company_name,
        "quarter": format_quarter_cell(summary.quarter, summary.as_of_date),
        "what_happened": format_what_happened(summary.what_happened),
        "positives": format_list(summary.positives),
        "negatives": format_list(summary.negatives),
        "document_only_confidence_score": str(summary.document_only_confidence_score),
        "confidence_score": str(summary.confidence_score),
        "analysis": format_analysis_csv(summary.analysis),
    }


def summary_to_excel_row(summary: QuarterSummary) -> dict[str, str]:
    return {
        "Summary Type": summary.summary_type.title(),
        "Company Name": summary.company_name,
        "Quarter": format_quarter_cell(summary.quarter, summary.as_of_date),
        "What Happened": format_bullets(summary.what_happened),
        "Positives": format_bullets(summary.positives),
        "Negatives": format_bullets(summary.negatives),
        "Document-Only Score": str(summary.document_only_confidence_score),
        "Confidence Score": str(summary.confidence_score),
        "Analysis": format_analysis_bullets(summary.analysis),
    }


def sanitize_sheet_title(title: str, existing_titles: set[str]) -> str:
    sanitized = re.sub(r"[\[\]\:\*\?\/\\]", "", title).strip()
    if not sanitized:
        sanitized = "Company"
    sanitized = sanitized[:31]

    candidate = sanitized
    counter = 2
    while candidate in existing_titles:
        suffix = f" {counter}"
        candidate = f"{sanitized[:31 - len(suffix)]}{suffix}"
        counter += 1
    existing_titles.add(candidate)
    return candidate


def estimate_wrapped_line_count(value: object, column_width: float) -> int:
    text = str(value or "")
    if not text:
        return 1

    chars_per_line = max(8, int(column_width))
    wrapped_lines = 0
    for line in text.splitlines() or [""]:
        wrapped_lines += max(1, math.ceil(len(line) / chars_per_line))
    return wrapped_lines


def max_lines_at_max_height() -> int:
    return int(
        (EXCEL_MAX_ROW_HEIGHT - ROW_HEIGHT_PADDING) // POINTS_PER_WRAPPED_LINE
    )


def min_column_width_for_text(
    text: str,
    max_lines: int,
    min_width: int = 8,
    max_width: int = EXCEL_MAX_COLUMN_WIDTH,
) -> float:
    if not text.strip():
        return float(min_width)

    if estimate_wrapped_line_count(text, max_width) > max_lines:
        return float(max_width)

    low = min_width
    high = max_width
    while low < high:
        mid = (low + high) // 2
        if estimate_wrapped_line_count(text, mid) <= max_lines:
            high = mid
        else:
            low = mid + 1
    return float(low)


def compute_analysis_column_width(worksheet, analysis_column_index: int) -> float:
    default_width = EXCEL_COLUMN_WIDTHS["Analysis"]
    if worksheet.max_row < 2:
        return default_width

    max_lines = max_lines_at_max_height()
    required_width = default_width
    for row_index in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row_index, column=analysis_column_index).value
        if cell_value is None:
            continue
        required_width = max(
            required_width,
            min_column_width_for_text(str(cell_value), max_lines),
        )
    return min(required_width, float(EXCEL_MAX_COLUMN_WIDTH))


def estimate_row_height(worksheet, row_index: int, table_column_count: int) -> float:
    max_lines = 1
    for column_index in range(1, table_column_count + 1):
        column_letter = get_column_letter(column_index)
        column_width = worksheet.column_dimensions[column_letter].width or 12
        cell_value = worksheet.cell(row=row_index, column=column_index).value
        max_lines = max(
            max_lines,
            estimate_wrapped_line_count(cell_value, column_width),
        )

    estimated_height = (max_lines * POINTS_PER_WRAPPED_LINE) + ROW_HEIGHT_PADDING
    return min(
        max(float(MIN_DATA_ROW_HEIGHT), estimated_height),
        EXCEL_MAX_ROW_HEIGHT,
    )


def write_csv(rows: Sequence[QuarterSummary], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(summary_to_row(row))


def populate_excel_sheet(worksheet, rows: Sequence[QuarterSummary]) -> None:
    headers = [DISPLAY_HEADERS[column] for column in CSV_COLUMNS]
    worksheet.append(headers)

    for summary in rows:
        excel_row = summary_to_excel_row(summary)
        worksheet.append([excel_row[header] for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)
    centered_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    centered_columns = {
        "Summary Type",
        "Company Name",
        "Quarter",
        "Document-Only Score",
        "Confidence Score",
    }
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            header = worksheet.cell(row=1, column=cell.column).value
            cell.alignment = (
                centered_alignment if header in centered_columns else body_alignment
            )

    analysis_column_index = headers.index("Analysis") + 1
    for column_index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(column_index)
        if header == "Analysis":
            continue
        worksheet.column_dimensions[column_letter].width = EXCEL_COLUMN_WIDTHS[header]

    analysis_letter = get_column_letter(analysis_column_index)
    worksheet.column_dimensions[analysis_letter].width = compute_analysis_column_width(
        worksheet,
        analysis_column_index,
    )

    worksheet.row_dimensions[1].height = 28
    last_data_row = worksheet.max_row
    for row_index in range(2, last_data_row + 1):
        worksheet.row_dimensions[row_index].height = estimate_row_height(
            worksheet,
            row_index,
            TABLE_COLUMN_COUNT,
        )

    write_confidence_reference_key(
        worksheet,
        start_column=REFERENCE_KEY_START_COLUMN,
        table_last_row=last_data_row,
        spacer_column=REFERENCE_KEY_SPACER_COLUMN,
    )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(TABLE_COLUMN_COUNT)}{last_data_row}"
    )


def sort_quarter_summaries(rows: Sequence[QuarterSummary]) -> list[QuarterSummary]:
    return sorted(
        rows,
        key=lambda row: (
            row.company_name.lower(),
            *quarter_sort_key(row.quarter),
        ),
    )


def group_rows_by_company(rows: Sequence[QuarterSummary]) -> dict[str, list[QuarterSummary]]:
    grouped: dict[str, list[QuarterSummary]] = {}
    for row in sort_quarter_summaries(rows):
        grouped.setdefault(row.company_name, []).append(row)
    return grouped


def write_excel(
    rows: Sequence[QuarterSummary],
    output_path: Path,
    *,
    single_sheet: bool = False,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    sorted_rows = sort_quarter_summaries(rows)
    if single_sheet:
        worksheet = workbook.create_sheet("Earnings Summary")
        populate_excel_sheet(worksheet, sorted_rows)
    else:
        existing_titles: set[str] = set()
        grouped_rows = group_rows_by_company(rows)
        if not grouped_rows:
            worksheet = workbook.create_sheet("Earnings Summary")
            populate_excel_sheet(worksheet, [])
        else:
            for company_name, company_rows in grouped_rows.items():
                sheet_title = sanitize_sheet_title(company_name, existing_titles)
                worksheet = workbook.create_sheet(sheet_title)
                populate_excel_sheet(worksheet, company_rows)

    workbook.save(output_path)


def write_output(
    rows: Sequence[QuarterSummary],
    output_path: Path,
    *,
    single_sheet: bool = False,
) -> None:
    if output_path.suffix.lower() == ".xlsx":
        write_excel(rows, output_path, single_sheet=single_sheet)
        return
    write_csv(rows, output_path)

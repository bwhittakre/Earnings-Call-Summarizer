from __future__ import annotations

from functools import lru_cache
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REFERENCE_KEY_TITLE = "Confidence Score Reference"
REFERENCE_KEY_TEXT_PATH = Path(__file__).with_name("confidence_reference_key.txt")
REFERENCE_KEY_COLUMN_COUNT = 4
REFERENCE_KEY_COLUMN_WIDTH = 20
REFERENCE_KEY_SPACER_WIDTH = 2.5
REFERENCE_KEY_BODY_ROW = 2
REFERENCE_KEY_END_ROW = 47

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SIDEBAR_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)


@lru_cache(maxsize=1)
def load_reference_key_text() -> str:
    return REFERENCE_KEY_TEXT_PATH.read_text(encoding="utf-8").strip()


def write_confidence_reference_key(
    worksheet,
    *,
    start_column: int,
    table_last_row: int,
    spacer_column: int | None = None,
) -> None:
    if spacer_column is not None:
        spacer_letter = get_column_letter(spacer_column)
        worksheet.column_dimensions[spacer_letter].width = REFERENCE_KEY_SPACER_WIDTH

    end_column = start_column + REFERENCE_KEY_COLUMN_COUNT - 1
    start_letter = get_column_letter(start_column)
    end_letter = get_column_letter(end_column)

    for column_index in range(start_column, end_column + 1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = REFERENCE_KEY_COLUMN_WIDTH

    title_range = f"{start_letter}1:{end_letter}1"
    body_range = (
        f"{start_letter}{REFERENCE_KEY_BODY_ROW}:"
        f"{end_letter}{REFERENCE_KEY_END_ROW}"
    )
    worksheet.merge_cells(title_range)
    worksheet.merge_cells(body_range)

    title_cell = worksheet[f"{start_letter}1"]
    title_cell.value = REFERENCE_KEY_TITLE
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(color="FFFFFF", bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.border = THIN_BORDER

    body_cell = worksheet[f"{start_letter}{REFERENCE_KEY_BODY_ROW}"]
    body_cell.value = load_reference_key_text()
    body_cell.fill = SIDEBAR_FILL
    body_cell.font = Font(size=10)
    body_cell.alignment = Alignment(vertical="top", wrap_text=True)
    body_cell.border = THIN_BORDER

    sidebar_end_row = max(table_last_row, REFERENCE_KEY_END_ROW)
    for row_index in range(1, sidebar_end_row + 1):
        for column_index in range(start_column, end_column + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if row_index == 1:
                continue
            if row_index == REFERENCE_KEY_BODY_ROW and column_index == start_column:
                continue
            cell.fill = SIDEBAR_FILL
            cell.border = THIN_BORDER

    worksheet.row_dimensions[1].height = max(
        worksheet.row_dimensions[1].height or 0,
        28,
    )

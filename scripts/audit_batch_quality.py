#!/usr/bin/env python3
"""Audit batch Excel output for blank positives/negatives and evidence drop rates."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from src.paths import EVIDENCE_AUDIT_DIR
from src.scoring.analysis_score import parse_analysis_weight

DEFAULT_MAX_DROP_RATE = 0.85
DEFAULT_MIN_VALIDATED_FACTORS = 1


def _parse_quarter_cell(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text.splitlines()[0].strip()


def _find_header_row(sheet) -> tuple[int, dict[str, int]]:
    for header_row in (1, 2):
        headers = [cell.value for cell in sheet[header_row]]
        if headers and "Quarter" in headers:
            col = {header: index + 1 for index, header in enumerate(headers) if header}
            return header_row, col
    raise ValueError("Could not locate header row with Quarter column")


def _latest_audit_for_label(label_prefix: str) -> dict | None:
    pattern = re.compile(re.escape(label_prefix.replace("/", "_")) + r".*\.json$")
    candidates = sorted(
        (
            path
            for path in EVIDENCE_AUDIT_DIR.glob("*.json")
            if pattern.search(path.name)
        ),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        return None
    return json.loads(candidates[0].read_text(encoding="utf-8"))


def _parse_fetch_summary(value: object) -> dict[str, object]:
    text = str(value or "").strip()
    docs_part = text.split(";")[0].strip() if text else ""
    doc_types = [part.strip() for part in docs_part.split("+")] if docs_part else []
    doc_types = [part for part in doc_types if part and part != "none"]
    transcript_found = "Transcript(found)" in text
    transcript_missing = "Transcript(missing)" in text
    trimmed = "trimmed" in text.lower()
    return {
        "docs_fetched": len(doc_types),
        "doc_types": doc_types,
        "transcript_available": transcript_found and not transcript_missing,
        "corpus_trimmed": trimmed,
    }


def _audit_metrics(audit: dict | None) -> dict[str, int]:
    if not audit:
        return {
            "evidence_kept": 0,
            "evidence_anchored": 0,
            "evidence_rescued": 0,
            "evidence_dropped": 0,
            "dropped_positives": 0,
            "dropped_negatives": 0,
            "dropped_analysis": 0,
            "dropped_what_happened": 0,
        }

    dropped_by_field: dict[str, int] = {}
    for entry in audit.get("dropped", []):
        field = entry.get("field", "unknown")
        dropped_by_field[field] = dropped_by_field.get(field, 0) + 1

    auto_anchored = len(audit.get("auto_anchored", []))
    rescued = len(audit.get("rescued", []))
    dropped = len(audit.get("dropped", []))
    kept = audit.get("verbatim_kept", audit.get("kept_count", 0))

    return {
        "evidence_kept": int(kept or 0),
        "evidence_anchored": auto_anchored,
        "evidence_rescued": rescued,
        "evidence_dropped": dropped,
        "dropped_positives": dropped_by_field.get("positives", 0),
        "dropped_negatives": dropped_by_field.get("negatives", 0),
        "dropped_analysis": dropped_by_field.get("analysis", 0),
        "dropped_what_happened": dropped_by_field.get("what_happened", 0),
    }


def _validated_factor_count(workbook, quarter: str) -> int:
    if "Validated Factors" not in workbook.sheetnames:
        return 0
    sheet = workbook["Validated Factors"]
    header_row = 2 if "Confidence Score uses Edgar" in str(sheet.cell(1, 1).value or "") else 1
    count = 0
    for row_index in range(header_row + 1, sheet.max_row + 1):
        cell_quarter = _parse_quarter_cell(sheet.cell(row_index, 1).value)
        if cell_quarter == quarter:
            count += 1
    return count


def audit_workbook(workbook_path: Path, output_csv: Path | None) -> list[dict]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = (
        workbook["Batch Backtest"]
        if "Batch Backtest" in workbook.sheetnames
        else workbook.active
    )
    header_row, col = _find_header_row(sheet)

    rows: list[dict] = []
    for row_index in range(header_row + 1, sheet.max_row + 1):
        quarter = _parse_quarter_cell(sheet.cell(row_index, col["Quarter"]).value)
        positives = sheet.cell(row_index, col["Positives"]).value
        negatives = sheet.cell(row_index, col["Negatives"]).value
        analysis = sheet.cell(row_index, col["Analysis"]).value
        confidence = sheet.cell(row_index, col["Confidence Score"]).value
        fetch_summary = (
            sheet.cell(row_index, col["Fetch Summary"]).value
            if "Fetch Summary" in col
            else ""
        )

        pos_blank = not positives or not str(positives).strip()
        neg_blank = not negatives or not str(negatives).strip()
        analysis_text = str(analysis or "")
        weighted_bullets = sum(
            1
            for line in analysis_text.splitlines()
            if parse_analysis_weight(line.lstrip("• ").split(" — ")[0]) is not None
        )
        analysis_thin = weighted_bullets < 4

        audit = _latest_audit_for_label(f"{quarter}") or _latest_audit_for_label(
            quarter.replace("-", "_")
        )
        metrics = _audit_metrics(audit)
        fetch_meta = _parse_fetch_summary(fetch_summary)
        validated_factor_count = _validated_factor_count(workbook, quarter)

        total_claims = metrics["evidence_kept"] + metrics["evidence_dropped"]
        drop_rate = (
            metrics["evidence_dropped"] / total_claims if total_claims else 0.0
        )

        rows.append(
            {
                "quarter": quarter,
                "confidence_score": confidence,
                "blank_positives": pos_blank,
                "blank_negatives": neg_blank,
                "analysis_weighted_bullets": weighted_bullets,
                "thin_analysis": analysis_thin,
                "fetch_summary": str(fetch_summary or ""),
                "docs_fetched": fetch_meta["docs_fetched"],
                "doc_types": ",".join(fetch_meta["doc_types"]),
                "corpus_trimmed": fetch_meta["corpus_trimmed"],
                "transcript_available": fetch_meta["transcript_available"],
                "evidence_kept": metrics["evidence_kept"],
                "evidence_anchored": metrics["evidence_anchored"],
                "evidence_rescued": metrics["evidence_rescued"],
                "evidence_dropped": metrics["evidence_dropped"],
                "drop_rate": round(drop_rate, 3),
                "dropped_positives": metrics["dropped_positives"],
                "dropped_negatives": metrics["dropped_negatives"],
                "dropped_analysis": metrics["dropped_analysis"],
                "dropped_what_happened": metrics["dropped_what_happened"],
                "validated_factor_count": validated_factor_count,
                "backfilled_from_analysis": ",".join(
                    audit.get("backfilled_from_analysis", [])
                )
                if audit
                else "",
            }
        )

    if output_csv and rows:
        output_csv.parent.mkdir(parents=True, exist_ok=True)
        with output_csv.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)

    return rows


def _check_regression(
    rows: list[dict],
    *,
    max_drop_rate: float,
    min_validated_factors: int,
) -> list[str]:
    failures: list[str] = []
    for row in rows:
        quarter = row["quarter"]
        if not quarter or row.get("confidence_score") in ("", None):
            continue
        if row["drop_rate"] > max_drop_rate and row["evidence_dropped"] > 0:
            failures.append(
                f"{quarter}: drop_rate {row['drop_rate']:.1%} exceeds {max_drop_rate:.0%}"
            )
        if row["validated_factor_count"] < min_validated_factors:
            failures.append(
                f"{quarter}: validated_factor_count {row['validated_factor_count']} "
                f"< {min_validated_factors}"
            )
    return failures


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit batch Excel quality metrics")
    parser.add_argument("workbook", type=Path, help="Path to batch .xlsx file")
    parser.add_argument(
        "--output-csv",
        type=Path,
        help="Optional CSV summary output path",
    )
    parser.add_argument(
        "--fail-on-regression",
        action="store_true",
        help="Exit non-zero when drop rate or validated factor counts exceed thresholds",
    )
    parser.add_argument(
        "--max-drop-rate",
        type=float,
        default=DEFAULT_MAX_DROP_RATE,
        help=f"Maximum allowed evidence drop rate (default: {DEFAULT_MAX_DROP_RATE})",
    )
    parser.add_argument(
        "--min-validated-factors",
        type=int,
        default=DEFAULT_MIN_VALIDATED_FACTORS,
        help=f"Minimum validated factors per scored quarter (default: {DEFAULT_MIN_VALIDATED_FACTORS})",
    )
    args = parser.parse_args()

    rows = audit_workbook(args.workbook, args.output_csv)
    total = len(rows)
    blank_pos = sum(1 for row in rows if row["blank_positives"])
    blank_neg = sum(1 for row in rows if row["blank_negatives"])
    thin = sum(1 for row in rows if row["thin_analysis"])
    avg_factors = (
        sum(row["validated_factor_count"] for row in rows) / total if total else 0
    )

    print(f"Quarters audited: {total}")
    print(f"Blank positives: {blank_pos}/{total}")
    print(f"Blank negatives: {blank_neg}/{total}")
    print(f"Thin analysis (<4 weighted bullets): {thin}/{total}")
    print(f"Avg validated factors per quarter: {avg_factors:.1f}")
    if args.output_csv:
        print(f"Wrote {args.output_csv}")

    if args.fail_on_regression:
        failures = _check_regression(
            rows,
            max_drop_rate=args.max_drop_rate,
            min_validated_factors=args.min_validated_factors,
        )
        if failures:
            print("Regression checks failed:", file=sys.stderr)
            for failure in failures:
                print(f"  - {failure}", file=sys.stderr)
            return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
